from datetime import timedelta
import logging
import os
from flask import Flask, request, render_template
from werkzeug.middleware.proxy_fix import ProxyFix
from flask_compress import Compress
from flask_minify import minify
from apscheduler.schedulers.background import BackgroundScheduler
import io
from flask import render_template
from openpyxl import load_workbook

app = Flask(__name__)
app.app_context()
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
compress = Compress()
compress.init_app(app)
# dev="development" 
dev="production"
ports = '8080'

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == "POST":
        return voucher(request)
    else:
        return render_template('index.html')

ALLOWED_EXTENSIONS_EXCEL = {'xls', 'xlsx'}
ALLOWED_EXTENSIONS_IMAGE = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

def parse_excel(file):
    workbook = load_workbook(filename=file, data_only=True)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append([cell for cell in row])
    return data

def create_vouch(data, configs, templs, logo_path=None):

    if templs == 1:
        return render_template('A4.html', data=data, configs=configs, logo_path=logo_path)
    elif templs == 2:
        return render_template('thermal.html', data=data, configs=configs, logo_path=logo_path)
    elif templs == 3:
        return render_template('thermal_cafe.html', data=data, configs=configs, logo_path=logo_path)
    else:
        return 'invalid templates'

def voucher(request):
    if request.method == 'POST':
        if 'excel' not in request.files or request.files['excel'].filename == '':
            return "No Excel file uploaded", 400

        excel_file = request.files['excel']
        if allowed_file(excel_file.filename, {'xls', 'xlsx'}):
            excel_data = parse_excel(excel_file)
            excel_data_without_first_element = excel_data[1:]

            configs = request.form.getlist('config')

            templs = int(request.form['templates'])

            logo_path = None
            logo_set = None
            if 'logo' in request.files and allowed_file(request.files['logo'].filename, {'png', 'jpg', 'jpeg'}):
                logo_file = request.files['logo']
                logo_path = f'{app.root_path}/static/tmp/{logo_file.filename}'
                logo_file.save(logo_path)
                logo_set = f'/static/tmp/{logo_file.filename}'
            if 'namevoucher' in request.form:
                namevoucher = request.form['namevoucher']
                if namevoucher != '':
                    configs.append(f'namevoucher: {namevoucher}')
    
            response = create_vouch(excel_data_without_first_element, configs, templs, logo_set)
            
            return response
        else:
            return "Invalid file type for Excel file", 400
        

def cron10():
    pics = [file for file in os.listdir(f'{app.root_path}/static/tmp') if os.path.isfile(os.path.join(f'{app.root_path}/static/tmp', file))]
    for entry in pics:
        os.remove(os.path.join(f"{app.root_path}/static/tmp/{entry}"))

if not app.debug or os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
    scheduler = BackgroundScheduler(daemon=True)
    scheduler.add_job(cron10, 'interval', minutes=1)
    scheduler.start()

if __name__ == "__main__":
    with app.app_context():
        app.wsgi_app = ProxyFix(app.wsgi_app)
        if dev == 'production':
            logging.basicConfig(level=logging.ERROR, format='%(levelname)s - %(message)s')
            minify(app=app)
            app.run(debug=False, threaded=True, use_reloader=False)
        else:
            logging.basicConfig(level=logging.DEBUG, format='%(levelname)s - %(message)s')
            minify(app=app, html=False, js=False, cssless=False)
            app.run(debug=True, threaded=True)
